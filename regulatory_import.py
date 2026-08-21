# -*- coding: utf-8 -*-
"""Importers for the controlled regulatory reference files.

One importer per source, each declaring the EXACT column signature it supports.
A file whose header does not match is rejected by name rather than parsed
loosely - see the REACH change request, clause 4.2: "If ECHA changes an export
structure, the importer shall reject the unsupported structure clearly rather
than silently mis-map columns."

That rule is why these are separate functions with hard-coded signatures rather
than one clever parser. A clever parser that copes with a changed layout is a
parser that will one day read the wrong column and report a compliance result
from it.

The original file is always retained. Nothing here is the source of truth; it
is a parse of a file that is, and every record carries the row it came from.
"""

import csv
import datetime as dt
import io
import re

import regulatory_reference as rr


class ImportRejected(Exception):
    """The file is not the file this importer supports. Carries a message
    written for the person who uploaded it, naming what was expected."""


# ---------------------------------------------------------------------------
# Annex VI to CLP - the harmonised classification table
# ---------------------------------------------------------------------------
# ECHA publishes this as an Excel file, one per Adaptation to Technical
# Progress, from
#   https://echa.europa.eu/information-on-chemicals/annex-vi-to-clp
# The file states on its own first row that it is unofficial and that the
# binding text is Table 3 to Annex VI in the Official Journal. That disclaimer
# is stored with the dataset rather than dropped, because a readiness report
# that leans on this table should say what it leaned on.

ANNEX_VI_SIGNATURE = (
    "Index No", "ATP", "CELEX", "Chemical Name", "EC No", "CAS No",
    "Hazard Class and Category Code(s)", "Classification Hazard Statement Code(s)",
    "Labelling Pictogram, Signal Word Code(s)", "Labelling Hazard Statement Code(s)",
    "Labelling Suppl. Hazard Statement Code(s)", "M, SCL, ATE", "Notes", "Comment",
    "In application", "EUR-Lex Link",
)
ANNEX_VI_HEADER_ROW = 3
ANNEX_VI_PARSER_VERSION = "v1"

_CAS_TOKEN = re.compile(r"\b\d{2,7}-\d{2}-\d\b")
_H_TOKEN = re.compile(r"\bH\d{3}[A-Za-z]{0,3}\b")


def _cas_numbers(value):
    """Every CAS-shaped token in a cell.

    The column is not one number per row. Real values include "7440-41-7", a
    bare "-" where none is allocated, and multi-substance entries such as
    "10043-35-3 [1]\\n11113-50-1 [2]" where the bracket refers back to a name
    on the same row. Each one becomes its own record, so a lookup by CAS finds
    the classification whichever of them the safety data sheet discloses."""
    return [m.group(0) for m in _CAS_TOKEN.finditer(str(value or ""))]


def _hazard_codes(value):
    """The hazard statement codes on a row, comma joined and de-duplicated."""
    seen, out = set(), []
    for m in _H_TOKEN.finditer(str(value or "")):
        c = m.group(0)
        if c not in seen:
            seen.add(c); out.append(c)
    return ",".join(out)


def _as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value:
        try:
            return dt.date.fromisoformat(str(value)[:10])
        except ValueError:
            return None
    return None


def parse_annex_vi(file_bytes):
    """(records, meta) from an ECHA Annex VI to CLP Excel export.

    Raises ImportRejected where the sheet or the header is not the supported
    one. Returns every substance, not only the prohibited ones: the criterion
    decides what is prohibited, the reference states what the classification
    is."""
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - openpyxl ships with the app
        raise ImportRejected("openpyxl is not available to read the Excel file.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ImportRejected(
            "This file could not be opened as an Excel workbook. Download the table again "
            "from %s and upload it unchanged." % rr.REFERENCE_SOURCES[rr.REFERENCE_HARMONISED_CLP]
        )

    sheet = next((n for n in wb.sheetnames if re.fullmatch(r"(CLP\d*|ATP\d+)", n or "")), None)
    if sheet is None:
        raise ImportRejected(
            "No harmonised classification sheet was found. Expected a sheet named for its "
            "adaptation, such as ATP23. This workbook has: %s." % ", ".join(wb.sheetnames)
        )

    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < ANNEX_VI_HEADER_ROW + 1:
        raise ImportRejected("Sheet %s has no data rows." % sheet)

    header = tuple((c or "").strip() if isinstance(c, str) else c
                   for c in rows[ANNEX_VI_HEADER_ROW - 1][:len(ANNEX_VI_SIGNATURE)])
    if header != ANNEX_VI_SIGNATURE:
        missing = [c for c in ANNEX_VI_SIGNATURE if c not in header]
        unexpected = [c for c in header if c and c not in ANNEX_VI_SIGNATURE]
        raise ImportRejected(
            "The column layout of sheet %s is not the one this importer supports (%s). "
            "Missing: %s. Unexpected: %s. Nothing has been read. If ECHA has changed the "
            "export, the importer needs updating rather than the file."
            % (sheet, ANNEX_VI_PARSER_VERSION, ", ".join(missing) or "none",
               ", ".join(unexpected) or "none")
        )

    disclaimer = rows[0][0] if rows and rows[0] else None
    records, skipped = [], 0
    for n, row in enumerate(rows[ANNEX_VI_HEADER_ROW:], start=ANNEX_VI_HEADER_ROW + 1):
        index_no = row[0]
        if not index_no:
            continue
        codes = _hazard_codes(row[7])
        cas_list = _cas_numbers(row[5])
        if not cas_list:
            # An entry with no allocated CAS - a group entry, or a substance
            # identified by index and EC number only. Kept out of the matching
            # set on purpose: this reference is looked up BY CAS, and a record
            # with none can never be found by it. Counted so the report can say
            # how many were set aside rather than pretending the file was
            # smaller than it is.
            skipped += 1
            continue
        for cas in cas_list:
            records.append({
                "cas_number": cas,
                "cas_normalised": rr.normalise_cas(cas),
                "ec_number": (str(row[4]).strip() if row[4] else None),
                "index_number": str(index_no).strip(),
                "substance_name": (str(row[3]).replace("\n", " ").strip() if row[3] else None),
                "classification_codes": codes,
                "entry_reference": (str(row[1]).strip() if row[1] else None),
                "in_application_date": _as_date(row[14]),
                "source_row_number": n,
            })
    meta = {
        "sheet": sheet,
        "version": sheet,
        "parser_name": "echa_annex_vi_clp",
        "parser_version": ANNEX_VI_PARSER_VERSION,
        "record_count": len(records),
        "entries_without_cas": skipped,
        "disclaimer": disclaimer,
    }
    return records, meta


# ---------------------------------------------------------------------------
# REACH Annex XVII Entry 43 - appendices 8 and 9
# ---------------------------------------------------------------------------
# Downloaded from ECHA as CSV:
#   https://echa.europa.eu/appendix-8-list-of-aromatic-amines
#   https://echa.europa.eu/appendix-9-list-of-azodyes
#
# Appendix 9 has exactly ONE entry, which looks like a truncated download and
# is not - verified against the consolidated Annex XVII text. Appendix 8 has
# twenty-two aromatic amines. Both are loaded into one reference set, each
# record tagged with the appendix it came from, because criterion 3.2 treats
# them differently: an azodye disclosed in a colourant is the dye itself, and a
# restricted aromatic amine disclosed in one is a stronger finding still.

ENTRY_43_SIGNATURE = ("Name", "EC Number", "Index Number", "CAS Number")
ENTRY_43_PARSER_VERSION = "v1"

APPENDIX_8 = "Entry 43 Appendix 8 - aromatic amines"
APPENDIX_9 = "Entry 43 Appendix 9 - azodyes"


def parse_entry_43(file_bytes, appendix):
    """(records, meta) from an ECHA Entry 43 appendix CSV.

    `appendix` is APPENDIX_8 or APPENDIX_9 and is recorded on every row - the
    file itself does not say which list it is."""
    if appendix not in (APPENDIX_8, APPENDIX_9):
        raise ImportRejected("Unknown appendix: %r" % appendix)
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        raise ImportRejected("The file is empty.")
    # A trailing comma on the header line produces a final empty column. That
    # is how ECHA's own export comes and is not a layout change.
    trimmed = tuple(h.strip() for h in header if h.strip())
    if trimmed != ENTRY_43_SIGNATURE:
        raise ImportRejected(
            "The column layout is not the one this importer supports (%s). Expected %s, "
            "found %s. Nothing has been read."
            % (ENTRY_43_PARSER_VERSION, ", ".join(ENTRY_43_SIGNATURE), ", ".join(trimmed) or "no columns")
        )

    records = []
    for n, row in enumerate(reader, start=2):
        if not any((c or "").strip() for c in row):
            continue
        name = (row[0] or "").strip()
        ec = (row[1] or "").strip() or None
        index_no = (row[2] or "").strip() or None
        cas_cell = (row[3] or "").strip() if len(row) > 3 else ""
        # Appendix 9's single entry leaves the CAS column empty and states the
        # number inside the name - "Component 1: CAS-No: 118685-33-9". Read it
        # from wherever it actually is rather than losing it.
        cas_list = _cas_numbers(cas_cell) or _cas_numbers(name)
        if not cas_list:
            raise ImportRejected(
                "Row %d has no CAS number in the CAS column or in the name, so it cannot be "
                "matched against a formulation. Row: %s" % (n, name[:120])
            )
        for cas in cas_list:
            records.append({
                "cas_number": cas,
                "cas_normalised": rr.normalise_cas(cas),
                "ec_number": ec,
                "index_number": index_no,
                "substance_name": re.sub(r"\s+", " ", name)[:400],
                "classification_codes": None,
                "entry_reference": appendix,
                "in_application_date": None,
                "source_row_number": n,
            })
    meta = {
        "version": appendix,
        "parser_name": "echa_entry_43",
        "parser_version": ENTRY_43_PARSER_VERSION,
        "record_count": len(records),
        "disclaimer": None,
    }
    return records, meta
